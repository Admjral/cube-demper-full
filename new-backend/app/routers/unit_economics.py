"""
Unit Economics Calculator API
Calculates margins, commissions, and profitability for Kaspi.kz products
"""

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from typing import Optional, List, Annotated
from decimal import Decimal
from datetime import datetime
from ..utils.security import escape_like
import re
import httpx
import csv
import io
import uuid
import asyncpg
import logging

logger = logging.getLogger(__name__)

from ..config import settings
from ..routers.auth import get_current_user
from ..core.database import get_db_pool

router = APIRouter()


# =============================================================================
# COMMISSION RATES DATA
# Based on Kaspi.kz tariffs 2026
# =============================================================================

# Categories with 6.4% commission (без НДС)
LOW_COMMISSION_CATEGORIES = {
    "Аптека": {
        "subcategories": ["Лекарства"],  # Only Лекарства has 6.4%, other Аптека is 10.9%
        "commission_no_vat": 6.4,
        "commission_with_vat": 7.3
    },
    "Продукты питания": {
        "subcategories": None,  # All subcategories
        "commission_no_vat": 6.4,
        "commission_with_vat": 7.3
    }
}

# Categories with 13.5% commission (без НДС)
HIGH_COMMISSION_CATEGORIES = {
    "Аксессуары": {
        "subcategories": [
            "Аксессуары для одежды и обуви",
            "Шапки, шарфы, перчатки",
            "Зонты, брелоки и портсигары",
            "Свадебные аксессуары",
            "Сумки, чемоданы, кошельки"  # Some items like Рюкзаки, Кошельки, Сумки
        ],
        "commission_no_vat": 13.5,
        "commission_with_vat": 15.5
    },
    "Телефоны и гаджеты": {
        "subcategories": ["Аксессуары для телефонов"],
        "commission_no_vat": 13.5,
        "commission_with_vat": 15.5
    },
    "ТВ, Аудио, Видео": {
        "subcategories": ["Чехлы для наушников"],  # Only some items
        "commission_no_vat": 13.5,
        "commission_with_vat": 15.5
    }
}

# Default commission rate for most categories
DEFAULT_COMMISSION = {
    "commission_no_vat": 10.9,
    "commission_with_vat": 12.5
}

# List of all top-level categories
ALL_CATEGORIES = [
    "Автотовары",
    "Аксессуары",
    "Аптека",
    "Бытовая техника",
    "Детские товары",
    "Досуг, книги",
    "Канцелярские товары",
    "Компьютеры",
    "Красота",
    "Мебель",
    "Обувь",
    "Одежда",
    "Подарки, сувениры",
    "Продукты питания",
    "Спорт, туризм",
    "Строительство, ремонт",
    "ТВ, Аудио, Видео",
    "Телефоны и гаджеты",
    "Товары для дома",
    "Украшения"
]


# =============================================================================
# DELIVERY TARIFFS DATA
# Based on Kaspi.kz delivery tariffs 2026
# =============================================================================

# Delivery tariffs for Kaspi Доставка (seller delivery to Kaspi points)
# НДС 16% added on top of base tariffs (Kaspi Guide shows prices without НДС)
DELIVERY_VAT_RATE = 0.16

# Tariffs from Kaspi Guide (с 1 января 2026, без НДС)
# For orders < 10,000 ₸ — flat rate by price (same for all delivery types)
# For orders >= 10,000 ₸ — by weight only (no price sub-ranges)
DELIVERY_TARIFFS = {
    "kaspi_city": {
        "name": "Kaspi Доставка (город)",
        "name_en": "Kaspi Delivery (city)",
        "description": "Доставка в пределах города",
        "tariffs": [
            {"min_price": 0, "max_price": 999, "cost": 49},
            {"min_price": 1000, "max_price": 2999, "cost": 149},
            {"min_price": 3000, "max_price": 4999, "cost": 199},
            {"min_price": 5000, "max_price": 9999, "cost": 699},
            {"min_price": 10000, "max_price": float('inf'), "weight_tariffs": [
                {"max_weight": 5, "cost": 1099},
                {"max_weight": 15, "cost": 1349},
                {"max_weight": 30, "cost": 2299},
                {"max_weight": 60, "cost": 2899},
                {"max_weight": 100, "cost": 4149},
                {"max_weight": 9999, "cost": 6449}
            ]}
        ]
    },
    "kaspi_kz": {
        "name": "Kaspi Доставка (по Казахстану)",
        "name_en": "Kaspi Delivery (nationwide)",
        "description": "Доставка по всему Казахстану",
        "tariffs": [
            {"min_price": 0, "max_price": 999, "cost": 49},
            {"min_price": 1000, "max_price": 2999, "cost": 149},
            {"min_price": 3000, "max_price": 4999, "cost": 199},
            {"min_price": 5000, "max_price": 9999, "cost": 799},
            {"min_price": 10000, "max_price": float('inf'), "weight_tariffs": [
                {"max_weight": 5, "cost": 1299},
                {"max_weight": 15, "cost": 1699},
                {"max_weight": 30, "cost": 3599},
                {"max_weight": 60, "cost": 5649},
                {"max_weight": 100, "cost": 8549},
                {"max_weight": 9999, "cost": 11999}
            ]}
        ]
    },
    "kaspi_express": {
        "name": "Kaspi Express",
        "name_en": "Kaspi Express",
        "description": "Экспресс-доставка в течение дня",
        "tariffs": [
            {"min_price": 0, "max_price": 999, "cost": 49},
            {"min_price": 1000, "max_price": 2999, "cost": 149},
            {"min_price": 3000, "max_price": 4999, "cost": 199},
            {"min_price": 5000, "max_price": 9999, "cost": 799},
            {"min_price": 10000, "max_price": float('inf'), "weight_tariffs": [
                {"max_weight": 5, "cost": 1699},
                {"max_weight": 15, "cost": 1849},
                {"max_weight": 30, "cost": 3149},
                {"max_weight": 60, "cost": 3599},
                {"max_weight": 100, "cost": 5599},
                {"max_weight": 9999, "cost": 8449}
            ]}
        ]
    },
    "self_pickup": {
        "name": "Самовывоз",
        "name_en": "Self Pickup",
        "description": "Покупатель забирает сам",
        "tariffs": [
            {"min_price": 0, "max_price": float('inf'), "cost": 0}
        ]
    }
}


# =============================================================================
# TAX REGIMES DATA
# =============================================================================

TAX_REGIMES = {
    "ip_simplified": {
        "name": "ИП упрощёнка",
        "name_en": "Individual entrepreneur (simplified)",
        "rate": 3.0,
        "description": "3% от дохода"
    },
    "ip_general": {
        "name": "ИП общеустановленный",
        "name_en": "Individual entrepreneur (general)",
        "rate": 10.0,
        "description": "10% ИПН"
    },
    "too_simplified": {
        "name": "ТОО упрощёнка",
        "name_en": "LLC (simplified)",
        "rate": 3.0,
        "description": "3% от дохода"
    },
    "too_general": {
        "name": "ТОО общеустановленный",
        "name_en": "LLC (general)",
        "rate": 20.0,
        "description": "20% КПН"
    },
    "patent": {
        "name": "Патент",
        "name_en": "Patent",
        "rate": 1.0,
        "description": "1% (фикс. платёж)"
    },
    "none": {
        "name": "Без налога",
        "name_en": "No tax",
        "rate": 0.0,
        "description": "Не учитывать налог"
    }
}


# =============================================================================
# PYDANTIC MODELS
# =============================================================================

class CategoryCommission(BaseModel):
    category: str
    subcategory: Optional[str] = None
    commission_no_vat: float
    commission_with_vat: float


class DeliveryCost(BaseModel):
    delivery_type: str
    name: str
    cost: float
    margin: float
    margin_percent: float
    profit: float


class CalculationRequest(BaseModel):
    selling_price: float = Field(..., gt=0, description="Цена продажи")
    purchase_price: float = Field(..., ge=0, description="Цена закупки")
    category: str = Field(..., description="Категория товара")
    subcategory: Optional[str] = Field(None, description="Подкатегория товара")
    weight_kg: float = Field(1.0, ge=0.1, le=31, description="Вес товара в кг")
    packaging_cost: float = Field(0, ge=0, description="Стоимость упаковки")
    other_costs: float = Field(0, ge=0, description="Прочие расходы")
    tax_regime: str = Field("ip_simplified", description="Налоговый режим")
    use_vat: bool = Field(False, description="Использовать комиссию с НДС")


class CalculationResult(BaseModel):
    selling_price: float
    purchase_price: float

    # Commission
    category: str
    commission_rate: float       # Base rate (e.g. 10.9%)
    commission_effective_rate: float  # Effective rate after НДС (e.g. 12.2%)
    commission_amount: float     # Actual amount deducted (with НДС)

    # Kaspi Pay
    kaspi_pay_rate: float        # 0.95%
    kaspi_pay_amount: float

    # Tax
    tax_regime: str
    tax_rate: float
    tax_amount: float

    # Other costs
    packaging_cost: float
    other_costs: float

    # Delivery scenarios
    delivery_scenarios: List[DeliveryCost]

    # Best scenario
    best_scenario: str
    best_profit: float
    best_margin: float


class ProductParseResult(BaseModel):
    product_name: Optional[str] = None
    price: Optional[float] = None
    category: Optional[str] = None
    subcategory: Optional[str] = None
    weight_kg: Optional[float] = None
    image_url: Optional[str] = None
    kaspi_url: str
    success: bool
    error: Optional[str] = None


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def get_commission_rate(category: str, subcategory: Optional[str] = None, use_vat: bool = False) -> tuple:
    """Get commission rate for a category/subcategory combination"""
    rate_key = "commission_with_vat" if use_vat else "commission_no_vat"

    # Check low commission categories (6.4%)
    if category in LOW_COMMISSION_CATEGORIES:
        cat_data = LOW_COMMISSION_CATEGORIES[category]
        if cat_data["subcategories"] is None:
            # All subcategories have low rate
            return cat_data[rate_key], cat_data["commission_no_vat"], cat_data["commission_with_vat"]
        elif subcategory and any(sub in subcategory for sub in cat_data["subcategories"]):
            return cat_data[rate_key], cat_data["commission_no_vat"], cat_data["commission_with_vat"]

    # Check high commission categories (13.5%)
    if category in HIGH_COMMISSION_CATEGORIES:
        cat_data = HIGH_COMMISSION_CATEGORIES[category]
        if cat_data["subcategories"] is None:
            return cat_data[rate_key], cat_data["commission_no_vat"], cat_data["commission_with_vat"]
        elif subcategory and any(sub in subcategory for sub in cat_data["subcategories"]):
            return cat_data[rate_key], cat_data["commission_no_vat"], cat_data["commission_with_vat"]

    # Default rate (10.9%)
    return DEFAULT_COMMISSION[rate_key], DEFAULT_COMMISSION["commission_no_vat"], DEFAULT_COMMISSION["commission_with_vat"]


def get_delivery_cost(delivery_type: str, price: float, weight_kg: float) -> float:
    """Calculate delivery cost based on type, price, and weight.
    Tariffs from Kaspi Guide are without НДС, we add 16% НДС on top.
    """
    if delivery_type not in DELIVERY_TARIFFS:
        return 0.0

    tariffs = DELIVERY_TARIFFS[delivery_type]["tariffs"]
    base_cost = 0.0

    for tariff in tariffs:
        if tariff["min_price"] <= price <= tariff.get("max_price", float('inf')):
            if "cost" in tariff:
                base_cost = tariff["cost"]
            elif "weight_tariffs" in tariff:
                for wt in tariff["weight_tariffs"]:
                    if weight_kg <= wt["max_weight"]:
                        base_cost = wt["cost"]
                        break
                else:
                    # If weight exceeds all tiers, use highest
                    base_cost = tariff["weight_tariffs"][-1]["cost"]
            break

    # Add НДС 16% on top (Kaspi Guide prices are without НДС)
    return round(base_cost * (1 + DELIVERY_VAT_RATE))


# =============================================================================
# API ENDPOINTS
# =============================================================================

@router.get("/categories")
async def get_categories():
    """Get list of all categories with their commission rates"""
    result = []
    for cat in ALL_CATEGORIES:
        rate, no_vat, with_vat = get_commission_rate(cat)
        result.append({
            "category": cat,
            "commission_no_vat": no_vat,
            "commission_with_vat": with_vat,
            "has_variable_rates": cat in HIGH_COMMISSION_CATEGORIES or cat in LOW_COMMISSION_CATEGORIES
        })
    return result


@router.get("/commission")
async def get_commission(
    category: str = Query(..., description="Category name"),
    subcategory: Optional[str] = Query(None, description="Subcategory name"),
    use_vat: bool = Query(False, description="Use VAT-included rate")
) -> CategoryCommission:
    """Get commission rate for a specific category"""
    rate, no_vat, with_vat = get_commission_rate(category, subcategory, use_vat)
    return CategoryCommission(
        category=category,
        subcategory=subcategory,
        commission_no_vat=no_vat,
        commission_with_vat=with_vat
    )


@router.get("/delivery-tariffs")
async def get_delivery_tariffs():
    """Get all delivery tariff structures"""
    result = {}
    for key, value in DELIVERY_TARIFFS.items():
        result[key] = {
            "name": value["name"],
            "name_en": value["name_en"],
            "description": value["description"]
        }
    return result


@router.get("/tax-regimes")
async def get_tax_regimes():
    """Get all available tax regimes"""
    return TAX_REGIMES


KASPI_PAY_RATE = 0.95  # Kaspi Pay fee: 0.95% of selling price
VAT_RATE = 12.0  # НДС rate in Kazakhstan: 12%


@router.post("/calculate", response_model=CalculationResult)
async def calculate_unit_economics(request: CalculationRequest):
    """
    Calculate full unit economics with all delivery scenarios
    """
    # Get base commission rate
    base_commission_rate, no_vat_rate, with_vat_rate = get_commission_rate(
        request.category,
        request.subcategory,
        request.use_vat
    )

    # Show base commission rate as-is (like AlgaTop).
    # Kaspi charges the base rate (e.g. 10.9%) from the selling price.
    # НДС handling is a separate accounting concern, not added on top.
    effective_commission_rate = no_vat_rate

    commission_amount = request.selling_price * (effective_commission_rate / 100)

    # Kaspi Pay fee (separate from marketplace commission)
    kaspi_pay_amount = request.selling_price * (KASPI_PAY_RATE / 100)

    # Get tax rate
    tax_data = TAX_REGIMES.get(request.tax_regime, TAX_REGIMES["none"])
    tax_rate = tax_data["rate"]
    tax_amount = request.selling_price * (tax_rate / 100)

    # Calculate base costs (without delivery)
    base_costs = (
        request.purchase_price +
        commission_amount +
        kaspi_pay_amount +
        tax_amount +
        request.packaging_cost +
        request.other_costs
    )

    # Calculate for each delivery scenario
    delivery_scenarios = []
    for delivery_type, delivery_data in DELIVERY_TARIFFS.items():
        delivery_cost = get_delivery_cost(
            delivery_type,
            request.selling_price,
            request.weight_kg
        )
        total_costs = base_costs + delivery_cost
        profit = request.selling_price - total_costs
        margin_percent = (profit / request.selling_price) * 100 if request.selling_price > 0 else 0

        delivery_scenarios.append(DeliveryCost(
            delivery_type=delivery_type,
            name=delivery_data["name"],
            cost=delivery_cost,
            margin=profit,
            margin_percent=round(margin_percent, 2),
            profit=profit
        ))

    # Sort by profit descending
    delivery_scenarios.sort(key=lambda x: x.profit, reverse=True)

    # Get best scenario
    best = delivery_scenarios[0] if delivery_scenarios else None

    return CalculationResult(
        selling_price=request.selling_price,
        purchase_price=request.purchase_price,
        category=request.category,
        commission_rate=base_commission_rate,
        commission_effective_rate=effective_commission_rate,
        commission_amount=round(commission_amount, 2),
        kaspi_pay_rate=KASPI_PAY_RATE,
        kaspi_pay_amount=round(kaspi_pay_amount, 2),
        tax_regime=tax_data["name"],
        tax_rate=tax_rate,
        tax_amount=round(tax_amount, 2),
        packaging_cost=request.packaging_cost,
        other_costs=request.other_costs,
        delivery_scenarios=delivery_scenarios,
        best_scenario=best.delivery_type if best else "self_pickup",
        best_profit=round(best.profit, 2) if best else 0,
        best_margin=round(best.margin_percent, 2) if best else 0
    )


# Category mapping from Kaspi URL slugs/keywords to our categories
CATEGORY_KEYWORDS = {
    "Бытовая техника": ["пылесос", "холодильник", "стиральн", "микроволнов", "кондиционер", "телевизор", "чайник", "утюг", "плита", "духов", "посудомо", "блендер", "миксер", "мультиварк", "кофемашин", "кофевар", "тостер", "гриль", "вытяжк", "морозильн", "сушильн"],
    "Телефоны и гаджеты": ["смартфон", "телефон", "iphone", "samsung galaxy", "xiaomi", "redmi", "poco", "realme", "honor", "huawei", "планшет", "ipad", "tablet", "smart watch", "умные часы", "наушник", "airpods", "buds", "power bank", "зарядк"],
    "Компьютеры": ["ноутбук", "laptop", "компьютер", "монитор", "клавиатур", "мышь", "мышка", "принтер", "сканер", "видеокарт", "процессор", "оперативн", "ssd", "hdd", "материнск"],
    "ТВ, Аудио, Видео": ["телевизор", "tv", "проектор", "колонк", "саундбар", "наушник", "микрофон", "видеокамер", "фотоаппарат"],
    "Красота": ["косметик", "парфюм", "духи", "шампун", "крем", "маска", "помад", "тушь", "тени", "пудр", "фен", "выпрямитель", "плойк", "эпилятор", "бритв", "триммер"],
    "Детские товары": ["детск", "коляск", "игрушк", "подгузник", "памперс", "кроватк", "манеж", "автокресл", "соск", "бутылочк"],
    "Одежда": ["платье", "рубашк", "футболк", "джинс", "брюк", "куртк", "пальто", "шуб", "костюм", "пиджак", "свитер", "кардиган", "боксер", "трусы", "белье", "нижне", "шорты", "леггинс", "носки", "майк", "толстовк", "худи"],
    "Обувь": ["кроссовк", "ботинк", "туфл", "сапог", "босоножк", "сандал", "кед", "слипон", "балетк"],
    "Спорт, туризм": ["велосипед", "тренажер", "гантел", "коврик", "палатк", "спальн", "рюкзак", "самокат", "скейт", "ролик", "лыж", "сноуборд", "бассейн"],
    "Мебель": ["диван", "кресл", "стол", "стул", "шкаф", "кровать", "матрас", "комод", "полк", "тумб"],
    "Товары для дома": ["посуд", "кастрюл", "сковород", "постельн", "полотенц", "шторы", "ковер", "светильник", "люстр", "лампа"],
    "Строительство, ремонт": ["дрель", "шуруповерт", "болгарк", "перфоратор", "краск", "обои", "плитк", "ламинат", "смесител"],
    "Автотовары": ["автомобил", "машин", "шин", "диск", "аккумулятор", "масл", "фильтр", "видеорегистратор", "навигатор"],
    "Аптека": ["витамин", "лекарств", "таблетк", "медицин", "тонометр", "ингалятор", "термометр", "бандаж"],
    "Продукты питания": ["продукт", "еда", "кофе", "чай", "шоколад", "конфет", "печень", "крупа", "макарон"],
    "Аксессуары": ["сумк", "кошелек", "часы", "очки", "ремень", "шарф", "перчатк", "шапк", "зонт"],
    "Украшения": ["кольц", "серьг", "браслет", "цепочк", "подвеск", "колье", "брошь"],
}


def detect_category_from_text(text: str) -> tuple:
    """Detect category from product name or description"""
    text_lower = text.lower()

    for category, keywords in CATEGORY_KEYWORDS.items():
        for keyword in keywords:
            if keyword in text_lower:
                return category, None

    return None, None


@router.get("/parse-url", response_model=ProductParseResult)
async def parse_kaspi_url(
    url: str = Query(..., description="Kaspi product URL"),
    current_user: dict = Depends(get_current_user),
):
    """
    Parse Kaspi product URL to extract product info and category

    Example URL: https://kaspi.kz/shop/p/samsung-galaxy-a54-8-256gb-chernyj-111730405/
    """
    # Validate URL - strict hostname check to prevent SSRF
    from urllib.parse import urlparse
    try:
        parsed = urlparse(url)
        hostname = (parsed.hostname or "").lower()
        if parsed.scheme not in ("http", "https") or not (hostname == "kaspi.kz" or hostname.endswith(".kaspi.kz")):
            return ProductParseResult(
                kaspi_url=url,
                success=False,
                error="Invalid Kaspi URL"
            )
    except Exception:
        return ProductParseResult(
            kaspi_url=url,
            success=False,
            error="Invalid Kaspi URL"
        )

    try:
        html = None

        # Try relay first (VPS → Railway → Kaspi) to avoid IP bans
        if settings.offers_relay_url and settings.offers_relay_secret:
            try:
                async with httpx.AsyncClient(timeout=20.0) as relay_client:
                    relay_resp = await relay_client.post(
                        f"{settings.offers_relay_url}/relay/parse-url",
                        json={"url": url},
                        headers={"Authorization": f"Bearer {settings.offers_relay_secret}"},
                    )
                    if relay_resp.status_code == 200:
                        data = relay_resp.json()
                        if data.get("status_code") == 200:
                            html = data["html"]
            except Exception as relay_err:
                logger.warning(f"Relay parse-url failed, falling back to direct: {relay_err}")

        # Fallback to direct request
        if html is None:
            async with httpx.AsyncClient(timeout=15.0) as client:
                headers = {
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
                    "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7"
                }
                response = await client.get(url, headers=headers, follow_redirects=True)

                if response.status_code != 200:
                    return ProductParseResult(
                        kaspi_url=url,
                        success=False,
                        error=f"Failed to fetch page: {response.status_code}"
                    )

                html = response.text

        if html is None:
            return ProductParseResult(
                kaspi_url=url,
                success=False,
                error="Failed to fetch page from both relay and direct"
            )

        # Extract product name from title
        product_name = None
        title_match = re.search(r'<title>([^<]+)</title>', html)
        if title_match:
            title = title_match.group(1)
            # Remove " — Kaspi.kz" suffix
            product_name = re.sub(r'\s*[—-]\s*Kaspi\.kz.*$', '', title).strip()

        # Extract price
        price = None
        # Try different price patterns
        price_patterns = [
            r'"price":\s*(\d+)',
            r'itemoffered-price["\s]*:\s*["\']?(\d+)',
            r'data-product-price="(\d+)"',
            r'"offers":\s*\{[^}]*"price":\s*(\d+)',
        ]
        for pattern in price_patterns:
            price_match = re.search(pattern, html)
            if price_match:
                price = float(price_match.group(1))
                break

        # Extract category
        category = None
        subcategory = None

        # Method 1: Try to find category in structured data
        cat_patterns = [
            r'"category":\s*"([^"]+)"',
            r'itemListElement[^}]*"name":\s*"([^"]+)"',
        ]
        for pattern in cat_patterns:
            cat_matches = re.findall(pattern, html)
            if cat_matches:
                for cat in cat_matches:
                    if cat in ALL_CATEGORIES:
                        category = cat
                        break
                # If no exact match, try keyword detection on structured category values
                if not category:
                    for cat in cat_matches:
                        detected, sub = detect_category_from_text(cat)
                        if detected:
                            category = detected
                            subcategory = sub
                            break
                if category:
                    break

        # Method 2: Detect from product name using keywords
        if not category and product_name:
            category, subcategory = detect_category_from_text(product_name)

        # Method 3: Detect from URL slug
        if not category:
            url_slug = url.lower()
            category, subcategory = detect_category_from_text(url_slug)

        # Method 4: Check breadcrumbs in HTML
        if not category:
            breadcrumb_match = re.search(r'breadcrumb[^>]*>([^<]+)', html, re.IGNORECASE)
            if breadcrumb_match:
                category, subcategory = detect_category_from_text(breadcrumb_match.group(1))

        # Extract image
        image_url = None
        img_match = re.search(r'"image":\s*"([^"]+)"', html)
        if img_match:
            image_url = img_match.group(1)

        # Extract weight from characteristics table
        weight_kg = None
        weight_patterns = [
            # "Вес" ... "0.23 кг" or "230 г" in specs table
            r'[Вв]ес[^<]*?</[^>]+>\s*<[^>]+>\s*([0-9.,]+)\s*(кг|г|kg|g)',
            # JSON-like: "weight": "0.23" or "Вес": "1.5 кг"
            r'"[Вв]ес":\s*"?([0-9.,]+)\s*(кг|г|kg|g)"?',
            # Spec row: Вес ... 1.5 кг
            r'[Вв]ес\s*(?:товара|брутто|нетто|,\s*кг)?\s*[:<]?\s*([0-9.,]+)\s*(кг|г|kg|g)',
            # "weight":"0.228" in JSON data
            r'"weight":\s*"?([0-9.,]+)"?',
        ]
        for pattern in weight_patterns:
            w_match = re.search(pattern, html)
            if w_match:
                try:
                    value = float(w_match.group(1).replace(',', '.'))
                    # Check if there's a unit group
                    unit = w_match.group(2).lower() if w_match.lastindex >= 2 else 'kg'
                    if unit in ('г', 'g'):
                        weight_kg = round(value / 1000, 3)
                    else:
                        weight_kg = round(value, 3)
                    # Sanity check: weight should be 0.01 - 100 kg
                    if weight_kg < 0.01 or weight_kg > 100:
                        weight_kg = None
                    else:
                        break
                except (ValueError, IndexError):
                    continue

        return ProductParseResult(
            product_name=product_name,
            price=price,
            category=category,
            subcategory=subcategory,
            weight_kg=weight_kg,
            image_url=image_url,
            kaspi_url=url,
            success=True
        )

    except httpx.TimeoutException:
        return ProductParseResult(
            kaspi_url=url,
            success=False,
            error="Request timeout"
        )
    except Exception as e:
        return ProductParseResult(
            kaspi_url=url,
            success=False,
            error=str(e)
        )


@router.get("/quick-calculate")
async def quick_calculate(
    selling_price: float = Query(..., gt=0, description="Selling price"),
    purchase_price: float = Query(..., ge=0, description="Purchase price"),
    category: str = Query("Автотовары", description="Category"),
    weight_kg: float = Query(1.0, ge=0.1, le=31, description="Weight in kg"),
    tax_regime: str = Query("ip_simplified", description="Tax regime"),
    use_vat: bool = Query(False, description="Use VAT-included commission")
):
    """
    Quick calculation without detailed breakdown
    Returns key metrics for each delivery type
    """
    request = CalculationRequest(
        selling_price=selling_price,
        purchase_price=purchase_price,
        category=category,
        weight_kg=weight_kg,
        tax_regime=tax_regime,
        use_vat=use_vat,
        packaging_cost=0,
        other_costs=0
    )

    result = await calculate_unit_economics(request)

    return {
        "commission_rate": result.commission_effective_rate,
        "commission": result.commission_amount,
        "kaspi_pay": result.kaspi_pay_amount,
        "tax_rate": result.tax_rate,
        "tax": result.tax_amount,
        "scenarios": {
            s.delivery_type: {
                "delivery_cost": s.cost,
                "profit": round(s.profit, 2),
                "margin": round(s.margin_percent, 2)
            }
            for s in result.delivery_scenarios
        },
        "best": {
            "type": result.best_scenario,
            "profit": result.best_profit,
            "margin": result.best_margin
        }
    }


# =============================================================================
# SAVED CALCULATIONS - PYDANTIC MODELS
# =============================================================================

class SaveCalculationRequest(BaseModel):
    name: str = Field(..., min_length=1, max_length=255, description="Product name")
    kaspi_url: Optional[str] = None
    image_url: Optional[str] = None
    selling_price: float = Field(..., gt=0)
    purchase_price: float = Field(..., ge=0)
    category: str
    subcategory: Optional[str] = None
    weight_kg: float = Field(1.0, ge=0.1, le=31)
    packaging_cost: float = Field(0, ge=0)
    other_costs: float = Field(0, ge=0)
    tax_regime: str = "ip_simplified"
    use_vat: bool = False
    notes: Optional[str] = None
    is_favorite: bool = False


class SavedCalculation(BaseModel):
    id: str
    name: str
    kaspi_url: Optional[str] = None
    image_url: Optional[str] = None
    selling_price: float
    purchase_price: float
    category: str
    subcategory: Optional[str] = None
    weight_kg: float
    packaging_cost: float
    other_costs: float
    tax_regime: str
    use_vat: bool
    commission_rate: Optional[float] = None
    commission_amount: Optional[float] = None
    tax_amount: Optional[float] = None
    best_scenario: Optional[str] = None
    best_profit: Optional[float] = None
    best_margin: Optional[float] = None
    notes: Optional[str] = None
    is_favorite: bool = False
    created_at: datetime
    updated_at: datetime


class SavedCalculationsList(BaseModel):
    items: List[SavedCalculation]
    total: int
    page: int
    page_size: int


# =============================================================================
# SAVED CALCULATIONS - CRUD ENDPOINTS
# =============================================================================

@router.post("/saved", response_model=SavedCalculation)
async def create_saved_calculation(
    request: SaveCalculationRequest,
    current_user: Annotated[dict, Depends(get_current_user)],
    pool: Annotated[asyncpg.Pool, Depends(get_db_pool)]
):
    """Save a calculation to the library"""
    user_id = current_user["id"]

    # Calculate results
    calc_request = CalculationRequest(
        selling_price=request.selling_price,
        purchase_price=request.purchase_price,
        category=request.category,
        subcategory=request.subcategory,
        weight_kg=request.weight_kg,
        packaging_cost=request.packaging_cost,
        other_costs=request.other_costs,
        tax_regime=request.tax_regime,
        use_vat=request.use_vat
    )
    result = await calculate_unit_economics(calc_request)

    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            INSERT INTO saved_calculations (
                user_id, name, kaspi_url, image_url,
                selling_price, purchase_price, category, subcategory,
                weight_kg, packaging_cost, other_costs, tax_regime, use_vat,
                commission_rate, commission_amount, tax_amount,
                best_scenario, best_profit, best_margin,
                notes, is_favorite
            ) VALUES (
                $1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13,
                $14, $15, $16, $17, $18, $19, $20, $21
            )
            RETURNING *
            """,
            user_id,
            request.name,
            request.kaspi_url,
            request.image_url,
            request.selling_price,
            request.purchase_price,
            request.category,
            request.subcategory,
            request.weight_kg,
            request.packaging_cost,
            request.other_costs,
            request.tax_regime,
            request.use_vat,
            result.commission_rate,
            result.commission_amount,
            result.tax_amount,
            result.best_scenario,
            result.best_profit,
            result.best_margin,
            request.notes,
            request.is_favorite
        )

    return SavedCalculation(
        id=str(row["id"]),
        name=row["name"],
        kaspi_url=row["kaspi_url"],
        image_url=row["image_url"],
        selling_price=float(row["selling_price"]),
        purchase_price=float(row["purchase_price"]),
        category=row["category"],
        subcategory=row["subcategory"],
        weight_kg=float(row["weight_kg"]),
        packaging_cost=float(row["packaging_cost"]),
        other_costs=float(row["other_costs"]),
        tax_regime=row["tax_regime"],
        use_vat=row["use_vat"],
        commission_rate=float(row["commission_rate"]) if row["commission_rate"] else None,
        commission_amount=float(row["commission_amount"]) if row["commission_amount"] else None,
        tax_amount=float(row["tax_amount"]) if row["tax_amount"] else None,
        best_scenario=row["best_scenario"],
        best_profit=float(row["best_profit"]) if row["best_profit"] else None,
        best_margin=float(row["best_margin"]) if row["best_margin"] else None,
        notes=row["notes"],
        is_favorite=row["is_favorite"],
        created_at=row["created_at"],
        updated_at=row["updated_at"]
    )


@router.get("/saved", response_model=SavedCalculationsList)
async def list_saved_calculations(
    current_user: Annotated[dict, Depends(get_current_user)],
    pool: Annotated[asyncpg.Pool, Depends(get_db_pool)],
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    favorites_only: bool = Query(False),
    search: Optional[str] = Query(None)
):
    """List saved calculations for the current user"""
    user_id = current_user["id"]
    offset = (page - 1) * page_size

    async with pool.acquire() as conn:
        # Build query
        where_clauses = ["user_id = $1"]
        params = [user_id]
        param_idx = 2

        if favorites_only:
            where_clauses.append("is_favorite = TRUE")

        if search:
            where_clauses.append(f"(name ILIKE ${param_idx} OR category ILIKE ${param_idx})")
            params.append(f"%{escape_like(search)}%")
            param_idx += 1

        where_sql = " AND ".join(where_clauses)

        # Get total
        total = await conn.fetchval(
            f"SELECT COUNT(*) FROM saved_calculations WHERE {where_sql}",
            *params
        )

        # Get items
        rows = await conn.fetch(
            f"""
            SELECT * FROM saved_calculations
            WHERE {where_sql}
            ORDER BY is_favorite DESC, updated_at DESC
            LIMIT ${param_idx} OFFSET ${param_idx + 1}
            """,
            *params, page_size, offset
        )

    items = [
        SavedCalculation(
            id=str(row["id"]),
            name=row["name"],
            kaspi_url=row["kaspi_url"],
            image_url=row["image_url"],
            selling_price=float(row["selling_price"]),
            purchase_price=float(row["purchase_price"]),
            category=row["category"],
            subcategory=row["subcategory"],
            weight_kg=float(row["weight_kg"]),
            packaging_cost=float(row["packaging_cost"]),
            other_costs=float(row["other_costs"]),
            tax_regime=row["tax_regime"],
            use_vat=row["use_vat"],
            commission_rate=float(row["commission_rate"]) if row["commission_rate"] else None,
            commission_amount=float(row["commission_amount"]) if row["commission_amount"] else None,
            tax_amount=float(row["tax_amount"]) if row["tax_amount"] else None,
            best_scenario=row["best_scenario"],
            best_profit=float(row["best_profit"]) if row["best_profit"] else None,
            best_margin=float(row["best_margin"]) if row["best_margin"] else None,
            notes=row["notes"],
            is_favorite=row["is_favorite"],
            created_at=row["created_at"],
            updated_at=row["updated_at"]
        )
        for row in rows
    ]

    return SavedCalculationsList(
        items=items,
        total=total,
        page=page,
        page_size=page_size
    )


@router.get("/saved/{calculation_id}", response_model=SavedCalculation)
async def get_saved_calculation(
    calculation_id: str,
    current_user: Annotated[dict, Depends(get_current_user)],
    pool: Annotated[asyncpg.Pool, Depends(get_db_pool)]
):
    """Get a specific saved calculation"""
    user_id = current_user["id"]

    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT * FROM saved_calculations WHERE id = $1 AND user_id = $2",
            uuid.UUID(calculation_id),
            user_id
        )

    if not row:
        raise HTTPException(status_code=404, detail="Calculation not found")

    return SavedCalculation(
        id=str(row["id"]),
        name=row["name"],
        kaspi_url=row["kaspi_url"],
        image_url=row["image_url"],
        selling_price=float(row["selling_price"]),
        purchase_price=float(row["purchase_price"]),
        category=row["category"],
        subcategory=row["subcategory"],
        weight_kg=float(row["weight_kg"]),
        packaging_cost=float(row["packaging_cost"]),
        other_costs=float(row["other_costs"]),
        tax_regime=row["tax_regime"],
        use_vat=row["use_vat"],
        commission_rate=float(row["commission_rate"]) if row["commission_rate"] else None,
        commission_amount=float(row["commission_amount"]) if row["commission_amount"] else None,
        tax_amount=float(row["tax_amount"]) if row["tax_amount"] else None,
        best_scenario=row["best_scenario"],
        best_profit=float(row["best_profit"]) if row["best_profit"] else None,
        best_margin=float(row["best_margin"]) if row["best_margin"] else None,
        notes=row["notes"],
        is_favorite=row["is_favorite"],
        created_at=row["created_at"],
        updated_at=row["updated_at"]
    )


@router.put("/saved/{calculation_id}", response_model=SavedCalculation)
async def update_saved_calculation(
    calculation_id: str,
    request: SaveCalculationRequest,
    current_user: Annotated[dict, Depends(get_current_user)],
    pool: Annotated[asyncpg.Pool, Depends(get_db_pool)]
):
    """Update a saved calculation"""
    user_id = current_user["id"]

    # Recalculate
    calc_request = CalculationRequest(
        selling_price=request.selling_price,
        purchase_price=request.purchase_price,
        category=request.category,
        subcategory=request.subcategory,
        weight_kg=request.weight_kg,
        packaging_cost=request.packaging_cost,
        other_costs=request.other_costs,
        tax_regime=request.tax_regime,
        use_vat=request.use_vat
    )
    result = await calculate_unit_economics(calc_request)

    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            UPDATE saved_calculations SET
                name = $3, kaspi_url = $4, image_url = $5,
                selling_price = $6, purchase_price = $7, category = $8, subcategory = $9,
                weight_kg = $10, packaging_cost = $11, other_costs = $12, tax_regime = $13, use_vat = $14,
                commission_rate = $15, commission_amount = $16, tax_amount = $17,
                best_scenario = $18, best_profit = $19, best_margin = $20,
                notes = $21, is_favorite = $22, updated_at = NOW()
            WHERE id = $1 AND user_id = $2
            RETURNING *
            """,
            uuid.UUID(calculation_id),
            user_id,
            request.name,
            request.kaspi_url,
            request.image_url,
            request.selling_price,
            request.purchase_price,
            request.category,
            request.subcategory,
            request.weight_kg,
            request.packaging_cost,
            request.other_costs,
            request.tax_regime,
            request.use_vat,
            result.commission_rate,
            result.commission_amount,
            result.tax_amount,
            result.best_scenario,
            result.best_profit,
            result.best_margin,
            request.notes,
            request.is_favorite
        )

    if not row:
        raise HTTPException(status_code=404, detail="Calculation not found")

    return SavedCalculation(
        id=str(row["id"]),
        name=row["name"],
        kaspi_url=row["kaspi_url"],
        image_url=row["image_url"],
        selling_price=float(row["selling_price"]),
        purchase_price=float(row["purchase_price"]),
        category=row["category"],
        subcategory=row["subcategory"],
        weight_kg=float(row["weight_kg"]),
        packaging_cost=float(row["packaging_cost"]),
        other_costs=float(row["other_costs"]),
        tax_regime=row["tax_regime"],
        use_vat=row["use_vat"],
        commission_rate=float(row["commission_rate"]) if row["commission_rate"] else None,
        commission_amount=float(row["commission_amount"]) if row["commission_amount"] else None,
        tax_amount=float(row["tax_amount"]) if row["tax_amount"] else None,
        best_scenario=row["best_scenario"],
        best_profit=float(row["best_profit"]) if row["best_profit"] else None,
        best_margin=float(row["best_margin"]) if row["best_margin"] else None,
        notes=row["notes"],
        is_favorite=row["is_favorite"],
        created_at=row["created_at"],
        updated_at=row["updated_at"]
    )


@router.patch("/saved/{calculation_id}/favorite")
async def toggle_favorite(
    calculation_id: str,
    current_user: Annotated[dict, Depends(get_current_user)],
    pool: Annotated[asyncpg.Pool, Depends(get_db_pool)]
):
    """Toggle favorite status"""
    user_id = current_user["id"]

    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            UPDATE saved_calculations
            SET is_favorite = NOT is_favorite, updated_at = NOW()
            WHERE id = $1 AND user_id = $2
            RETURNING id, is_favorite
            """,
            uuid.UUID(calculation_id),
            user_id
        )

    if not row:
        raise HTTPException(status_code=404, detail="Calculation not found")

    return {"id": str(row["id"]), "is_favorite": row["is_favorite"]}


@router.delete("/saved/{calculation_id}")
async def delete_saved_calculation(
    calculation_id: str,
    current_user: Annotated[dict, Depends(get_current_user)],
    pool: Annotated[asyncpg.Pool, Depends(get_db_pool)]
):
    """Delete a saved calculation"""
    user_id = current_user["id"]

    async with pool.acquire() as conn:
        result = await conn.execute(
            "DELETE FROM saved_calculations WHERE id = $1 AND user_id = $2",
            uuid.UUID(calculation_id),
            user_id
        )

    if result == "DELETE 0":
        raise HTTPException(status_code=404, detail="Calculation not found")

    return {"deleted": True}


# =============================================================================
# EXPORT ENDPOINTS
# =============================================================================

@router.get("/saved/export/csv")
async def export_to_csv(
    current_user: Annotated[dict, Depends(get_current_user)],
    pool: Annotated[asyncpg.Pool, Depends(get_db_pool)],
    favorites_only: bool = Query(False)
):
    """Export saved calculations to CSV"""
    user_id = current_user["id"]

    async with pool.acquire() as conn:
        where_clause = "user_id = $1"
        if favorites_only:
            where_clause += " AND is_favorite = TRUE"

        rows = await conn.fetch(
            f"""
            SELECT * FROM saved_calculations
            WHERE {where_clause}
            ORDER BY created_at DESC
            """,
            user_id
        )

    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow([
        "Название", "Категория", "Цена продажи", "Себестоимость",
        "Комиссия %", "Комиссия ₸", "Налог ₸",
        "Упаковка", "Прочие расходы", "Вес кг",
        "Лучший сценарий", "Прибыль", "Маржа %",
        "Налоговый режим", "Избранное", "Kaspi URL", "Заметки", "Дата создания"
    ])

    # Data rows
    for row in rows:
        scenario_names = {
            "self_pickup": "Самовывоз",
            "kaspi_city": "Kaspi город",
            "kaspi_kz": "Kaspi KZ",
            "kaspi_express": "Express"
        }
        writer.writerow([
            row["name"],
            row["category"],
            float(row["selling_price"]),
            float(row["purchase_price"]),
            float(row["commission_rate"]) if row["commission_rate"] else 0,
            float(row["commission_amount"]) if row["commission_amount"] else 0,
            float(row["tax_amount"]) if row["tax_amount"] else 0,
            float(row["packaging_cost"]),
            float(row["other_costs"]),
            float(row["weight_kg"]),
            scenario_names.get(row["best_scenario"], row["best_scenario"]),
            float(row["best_profit"]) if row["best_profit"] else 0,
            float(row["best_margin"]) if row["best_margin"] else 0,
            row["tax_regime"],
            "Да" if row["is_favorite"] else "Нет",
            row["kaspi_url"] or "",
            row["notes"] or "",
            row["created_at"].strftime("%Y-%m-%d %H:%M") if row["created_at"] else ""
        ])

    output.seek(0)

    # Return as downloadable file
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=unit_economics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        }
    )


@router.get("/saved/export/excel")
async def export_to_excel(
    current_user: Annotated[dict, Depends(get_current_user)],
    pool: Annotated[asyncpg.Pool, Depends(get_db_pool)],
    favorites_only: bool = Query(False)
):
    """Export saved calculations to Excel format"""
    user_id = current_user["id"]

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(
            status_code=501,
            detail="Excel export not available. Install openpyxl package."
        )

    async with pool.acquire() as conn:
        where_clause = "user_id = $1"
        if favorites_only:
            where_clause += " AND is_favorite = TRUE"

        rows = await conn.fetch(
            f"""
            SELECT * FROM saved_calculations
            WHERE {where_clause}
            ORDER BY created_at DESC
            """,
            user_id
        )

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Unit Economics"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        "Название", "Категория", "Цена продажи", "Себестоимость",
        "Комиссия %", "Комиссия ₸", "Налог ₸",
        "Упаковка", "Прочие расходы", "Вес кг",
        "Лучший сценарий", "Прибыль", "Маржа %",
        "Налоговый режим", "Избранное", "Дата создания"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Data rows
    scenario_names = {
        "self_pickup": "Самовывоз",
        "kaspi_city": "Kaspi город",
        "kaspi_kz": "Kaspi KZ",
        "kaspi_express": "Express"
    }

    for row_idx, row in enumerate(rows, 2):
        data = [
            row["name"],
            row["category"],
            float(row["selling_price"]),
            float(row["purchase_price"]),
            float(row["commission_rate"]) if row["commission_rate"] else 0,
            float(row["commission_amount"]) if row["commission_amount"] else 0,
            float(row["tax_amount"]) if row["tax_amount"] else 0,
            float(row["packaging_cost"]),
            float(row["other_costs"]),
            float(row["weight_kg"]),
            scenario_names.get(row["best_scenario"], row["best_scenario"]),
            float(row["best_profit"]) if row["best_profit"] else 0,
            float(row["best_margin"]) if row["best_margin"] else 0,
            row["tax_regime"],
            "Да" if row["is_favorite"] else "Нет",
            row["created_at"].strftime("%Y-%m-%d %H:%M") if row["created_at"] else ""
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border

            # Color profit cells
            if col == 12:  # Profit column
                if isinstance(value, (int, float)) and value > 0:
                    cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
                elif isinstance(value, (int, float)) and value < 0:
                    cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=unit_economics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        }
    )
